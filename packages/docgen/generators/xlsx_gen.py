"""Excel generator using openpyxl."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from packages.docgen.generators.base import DocumentContent, DocumentGenerator, DocumentSection

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14)


class XlsxGenerator(DocumentGenerator):
    """Generate .xlsx files from DocumentContent."""

    @property
    def mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return "xlsx"

    def generate(
        self, content: DocumentContent, template_path: str | None = None
    ) -> bytes:
        wb = Workbook()

        sections_with_tables = [s for s in content.sections if s.table]
        sections_without_tables = [s for s in content.sections if not s.table]

        if len(sections_with_tables) <= 1:
            # Everything in one sheet
            ws = wb.active
            ws.title = self._safe_sheet_name(content.title)
            self._write_single_sheet(ws, content)
        else:
            # Each table section gets its own sheet
            wb.remove(wb.active)
            for section in sections_with_tables:
                ws = wb.create_sheet(title=self._safe_sheet_name(section.heading))
                self._write_section_sheet(ws, section)

            # Non-table sections go on a summary sheet
            if sections_without_tables:
                ws = wb.create_sheet(title="Summary", index=0)
                row = 1
                cell = ws.cell(row=row, column=1, value=content.title)
                cell.font = _TITLE_FONT
                row += 2
                for section in sections_without_tables:
                    ws.cell(row=row, column=1, value=section.heading).font = Font(bold=True, size=12)
                    row += 1
                    if section.body:
                        for line in section.body.split("\n"):
                            stripped = line.strip()
                            if stripped:
                                ws.cell(row=row, column=1, value=stripped)
                                row += 1
                    row += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    def _write_single_sheet(self, ws, content: DocumentContent) -> None:
        row = 1
        cell = ws.cell(row=row, column=1, value=content.title)
        cell.font = _TITLE_FONT
        row += 2

        for section in content.sections:
            ws.cell(row=row, column=1, value=section.heading).font = Font(bold=True, size=12)
            row += 1

            if section.table:
                row = self._write_table(ws, section.table, row)
                row += 1
            elif section.body:
                for line in section.body.split("\n"):
                    stripped = line.strip()
                    if stripped:
                        ws.cell(row=row, column=1, value=stripped)
                        row += 1
            row += 1

    def _write_section_sheet(self, ws, section: DocumentSection) -> None:
        row = 1
        if section.body:
            for line in section.body.split("\n"):
                stripped = line.strip()
                if stripped:
                    ws.cell(row=row, column=1, value=stripped)
                    row += 1
            row += 1

        if section.table:
            self._write_table(ws, section.table, row)

    def _write_table(self, ws, table_data: list[list[str]], start_row: int) -> int:
        if not table_data:
            return start_row

        for r_idx, row_data in enumerate(table_data):
            for c_idx, cell_text in enumerate(row_data):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=cell_text)
                if r_idx == 0:
                    cell.fill = _HEADER_FILL
                    cell.font = _HEADER_FONT

        return start_row + len(table_data)

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """Truncate and sanitize sheet name to Excel's 31-char limit."""
        # Remove characters invalid in sheet names
        for ch in "[]:*?/\\":
            name = name.replace(ch, "")
        return name[:31] if name else "Sheet"
